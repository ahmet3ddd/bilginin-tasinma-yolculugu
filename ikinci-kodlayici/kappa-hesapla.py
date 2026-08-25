#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Aktarım Zinciri — kodlayıcılar arası uyum hesabı (Cohen kappa'sı).

Kullanım:
    python3 kappa-hesapla.py doldurulmus-form.xlsx

Gereken dosyalar, bu betikle aynı klasörde:
    CEVAP-ANAHTARI-gonullulere-verilmez.csv
    CEVAP-ANAHTARI-nesneler.csv

Çıktı: yüzde uyum, Cohen kappa'sı, halka bazında dökümanlar, ve anlaşmazlığa
düşülen satırların listesi. Hiçbir şeyi ağa göndermez.
"""
import csv, sys, os, collections

LINKS = ['paket','çözücü','bağlam','örtük bilgi','aparat','bakım']

def kappa(a, b):
    """Cohen's kappa for two equal-length lists of labels."""
    n = len(a)
    if n == 0: return float('nan'), 0.0, 0.0
    po = sum(1 for x, y in zip(a, b) if x == y) / n
    ca, cb = collections.Counter(a), collections.Counter(b)
    pe = sum((ca[k]/n) * (cb[k]/n) for k in set(ca) | set(cb))
    k = (po - pe) / (1 - pe) if pe != 1 else float('nan')
    return k, po, pe

def read_key(path, col):
    # utf-8-sig: dosyalar elektronik tablo uygulamaları doğru açsın diye BOM ile yazılır
    with open(path, newline='', encoding='utf-8-sig') as f:
        return [r[col] for r in csv.DictReader(f)]

def yorum(k):
    if k != k: return "hesaplanamadı"
    if k < 0.00: return "tesadüften kötü"
    if k < 0.20: return "çok zayıf"
    if k < 0.40: return "zayıf"
    if k < 0.60: return "orta"
    if k < 0.80: return "iyi (substantial)"
    return "çok iyi (almost perfect)"

def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl gerekiyor:  pip install openpyxl"); sys.exit(1)

    here = os.path.dirname(os.path.abspath(__file__))
    wb = load_workbook(sys.argv[1], data_only=True)

    # --- events: sheet "Olaylar", column F, from row 3 (row 2 is the worked example)
    we = wb['Olaylar']
    got = [(we.cell(row=r, column=6).value or '').strip()
           for r in range(3, we.max_row + 1)
           if we.cell(row=r, column=2).value]
    gold = read_key(os.path.join(here, 'CEVAP-ANAHTARI-gonullulere-verilmez.csv'), 'ilk_kodlama_TR')

    if len(got) != len(gold):
        print("UYARI: satır sayıları tutmuyor — form %d, anahtar %d" % (len(got), len(gold)))
    n = min(len(got), len(gold))
    pairs = [(g, h) for g, h in zip(gold[:n], got[:n]) if h]
    bos = n - len(pairs)
    a = [p[0] for p in pairs]; b = [p[1] for p in pairs]

    k, po, pe = kappa(a, b)
    print("=" * 62)
    print("OLAY KODLAMASI  —  %d satırın %d'i dolu%s" % (n, len(pairs), (", %d boş" % bos) if bos else ""))
    print("=" * 62)
    print("  yüzde uyum        : %.1f%%" % (po * 100))
    print("  beklenen uyum     : %.1f%%  (tesadüfen)" % (pe * 100))
    print("  Cohen kappa       : %.3f   → %s" % (k, yorum(k)))
    print()
    print("  halka bazında (ilk kodlama → gönüllü):")
    for L in LINKS:
        idx = [i for i, x in enumerate(a) if x == L]
        if not idx: continue
        agree = sum(1 for i in idx if b[i] == L)
        print("    %-12s %2d satır, %2d'inde aynı (%3.0f%%)" % (L, len(idx), agree, 100 * agree / len(idx)))
    print()
    dis = [(i + 1, a[i], b[i]) for i in range(len(a)) if a[i] != b[i]]
    print("  anlaşmazlık: %d satır" % len(dis))
    for i, x, y in dis:
        print("    satır %2d:  ilk=%-12s gönüllü=%-12s" % (i, x, y))

    # --- objects: sheet "Nesneler", column C, from row 3
    wo = wb['Nesneler']
    got2 = [(wo.cell(row=r, column=3).value or '').strip()
            for r in range(3, wo.max_row + 1) if wo.cell(row=r, column=1).value]
    gold2 = read_key(os.path.join(here, 'CEVAP-ANAHTARI-nesneler.csv'), 'ilk_kodlama_TR')
    names2 = read_key(os.path.join(here, 'CEVAP-ANAHTARI-nesneler.csv'), 'nesne')
    m = min(len(got2), len(gold2))
    pairs2 = [(g, h) for g, h in zip(gold2[:m], got2[:m]) if h]
    if pairs2:
        a2 = [p[0] for p in pairs2]; b2 = [p[1] for p in pairs2]
        k2, po2, _ = kappa(a2, b2)
        print()
        print("=" * 62)
        print("EN İNCE HALKA YARGISI  —  %d nesne" % len(pairs2))
        print("=" * 62)
        print("  yüzde uyum        : %.1f%%" % (po2 * 100))
        print("  Cohen kappa       : %.3f   → %s" % (k2, yorum(k2)))
        print("  (10 nesnede kappa çok gürültülüdür; yüzde uyumu rapor etmek daha dürüst.)")
        for i, (x, y) in enumerate(zip(a2, b2)):
            if x != y: print("    %-30s ilk=%-12s gönüllü=%-12s" % (names2[i][:30], x, y))

    print()
    print("Raporlanacak cümle örneği:")
    print('  "10 nesnelik bir alt örneklemde (79 olay) ikinci ve bağımsız bir kodlayıcı ile')
    print('   olay-halka kodlaması için Cohen kappa = %.2f (%%%.0f yüzde uyum) elde edilmiştir."' % (k, po * 100))

if __name__ == '__main__':
    main()
